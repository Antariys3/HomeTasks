import csv
import openpyxl

with open("workers_csv.csv", encoding="utf-8-sig") as file:
    f_csv = csv.reader(file)
    list_workers = [item[0].split(";") for item in f_csv]
print("Наш список для наглядности:")
print(list_workers)

work_book = openpyxl.Workbook()
work_book.create_sheet('Работники', index=0)
sheet = work_book['Работники']

index_ignore = -1
for row_index, row in enumerate(list_workers):
    cell_offset = 1
    for col_index, value in enumerate(row):
        # Проверка колонки на исключение и в index_ignore записываем индекс для игнорирования
        # вместо "Возраст" можно ввести любую переменную ('ID', 'Имя', 'Возраст') для исключения столбца
        # кроме 'Телефон' так как он последний в списке и его нечем затирать, но это легко исправить добавив
        # исключение continue или в конце нашего списка добавить пустой элемент. Но такого в задании нет.
        if value == "Возраст":
            index_ignore = col_index + 1
        elif index_ignore == col_index:
            cell_offset = 0  # смещение для игнорируемой колонки (затирание)
        cell = sheet.cell(row=row_index + 1, column=col_index + cell_offset)
        cell.value = value

work_book.save("workers_excel.xlsx")

# Дополнительное задание, проверять на второй странице этого же файла.
work_book.create_sheet('Доп. задание', index=1)
sheet_2 = work_book['Доп. задание']
list_person = [f"Person {i}" if i != 0 else "" for i in range(len(list_workers))]
print("list_person:", list_person)

for col_index, value in enumerate(list_person):
    row_index = 1
    cell = sheet_2.cell(row=row_index, column=col_index + 1)
    cell.value = value

index_ignore = -1
for row_index, row in enumerate(list_workers):
    cell_offset = 2
    for col_index, value in enumerate(row):
        if value == "Возраст":
            index_ignore = col_index + 1
        elif index_ignore == col_index:
            cell_offset = 1  # смещение для игнорируемой колонки (затирание)
        cell = sheet_2.cell(row=col_index + cell_offset, column=row_index + 1)
        cell.value = value

work_book.save("workers_excel.xlsx")
